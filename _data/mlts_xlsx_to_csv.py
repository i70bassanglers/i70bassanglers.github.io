#!/usr/bin/env python3
"""
Convert MLTS result sheets from the master spreadsheet to individual CSV files.

Each result sheet becomes one CSV in the output directory, named MM_DD_YYYY.csv.
Existing files are overwritten.

Usage:
    python3 mlts_xlsx_to_csv.py [xlsx_path] [output_dir]

Defaults:
    xlsx_path:  _data/temp_mlts.xlsx  (relative to this script)
    output_dir: _data/mlts_results/   (relative to this script)
"""

import csv
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")

SKIP_SHEETS = {"Points", "Team of the Year Payout"}

MONTHS = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

CSV_HEADER = ["angler1", "alt1", "angler2", "alt2", "fish", "alive", "big_bass", "weight", "penalty", "final_weight"]


def parse_date_from_title(title):
    """Extract a date from e.g. 'Milford Lake Team Series - Apr 25, 2026'."""
    m = re.search(r"(\w{3})\s+(\d{1,2}),\s+(\d{4})", str(title))
    if not m:
        return None
    month_name, day, year = m.groups()
    month = MONTHS.get(month_name)
    if not month:
        return None
    return datetime(int(year), month, int(day))


def parse_anglers(raw):
    """
    Parse 'Angler1/Angler2 (Alt)' into (angler1, alt1, angler2, alt2).
    The parenthetical, if present, is treated as the alternate for angler2.
    """
    raw = str(raw).strip()
    alt2 = "-"
    m = re.search(r"\(([^)]+)\)", raw)
    if m:
        alt2 = m.group(1).strip()
        raw = raw[: m.start()].strip()

    parts = raw.split("/", 1)
    angler1 = parts[0].strip()
    angler2 = parts[1].strip() if len(parts) > 1 else ""
    return angler1, "-", angler2, alt2


def fmt_float(value):
    """Format a float without trailing zeros past two decimal places."""
    if value is None:
        return ""
    f = float(value)
    # Round to 2 decimal places to avoid floating-point noise
    return f"{f:.2f}".rstrip("0").rstrip(".")


def is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() in ("", "-", " "))


def build_alternate_map(wb):
    """
    Parse the Points sheet to map (angler1_lower, alternate_lower) -> canonical_angler2.

    A row like 'Drew Easterday/Joe Doreen (Cole Miller)' means Cole Miller is a known
    alternate for Joe Doreen on Drew Easterday's team. When a result sheet lists
    'Drew Easterday/Cole Miller' with no parenthetical, we can substitute the canonical
    name (Joe Doreen) and record Cole Miller as alt2.
    """
    if "Points" not in wb.sheetnames:
        return {}

    alt_map = {}
    for row in wb["Points"].iter_rows(min_row=2, values_only=True):
        team_str = row[1]
        if is_blank(team_str):
            continue
        team_str = str(team_str).strip()
        m = re.search(r"\(([^)]+)\)", team_str)
        if not m:
            continue
        alternate = m.group(1).strip()
        base = team_str[: m.start()].strip()
        parts = base.split("/", 1)
        if len(parts) != 2:
            continue
        angler1 = parts[0].strip()
        angler2 = parts[1].strip()
        alt_map[(angler1.lower(), alternate.lower())] = (angler1, angler2, alternate)

    return alt_map


def convert_sheet(ws, output_dir, alt_map):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return

    title = rows[0][0]
    date = parse_date_from_title(title)
    if not date:
        print(f"  Skipping '{ws.title}': could not parse date from title '{title}'")
        return

    filename = date.strftime("%m_%d_%Y") + ".csv"
    filepath = os.path.join(output_dir, filename)

    out_rows = []
    notes = []
    for row in rows[2:]:  # skip title row and header row
        place = row[0]
        anglers = row[1]
        fish_count = row[2]
        alive_count = row[3]
        wgt = row[4]
        penalty = row[5]
        total = row[6]
        bb = row[7]

        # Skip the totals/summary row at the bottom (no place, no anglers)
        if is_blank(place) and is_blank(anglers):
            continue
        if is_blank(anglers):
            continue

        angler1, alt1, angler2, alt2 = parse_anglers(anglers)

        # If angler2 is a known alternate (no parenthetical in result sheet),
        # substitute the canonical team member and record who actually fished.
        if alt2 == "-":
            key = (angler1.lower(), angler2.lower())
            if key in alt_map:
                _, canonical_angler2, alternate = alt_map[key]
                notes.append(f"    alt sub: '{anglers}' — {angler2} → {canonical_angler2}")
                angler2 = canonical_angler2
                alt2 = alternate

        no_fish = (place == "-") or (
            isinstance(fish_count, (int, float)) and fish_count == 0
            and isinstance(total, (int, float)) and total == 0
        )

        fish_val = "-" if no_fish else (str(int(fish_count)) if fish_count is not None else "-")
        alive_val = "-" if no_fish else (str(int(alive_count)) if alive_count is not None else "-")
        bb_val = "-" if is_blank(bb) else fmt_float(bb)
        weight_val = fmt_float(wgt) if not is_blank(wgt) else "0"
        penalty_val = "" if is_blank(penalty) else fmt_float(penalty)
        final_val = fmt_float(total) if not is_blank(total) else "0"

        out_rows.append([angler1, alt1, angler2, alt2, fish_val, alive_val, bb_val, weight_val, penalty_val, final_val])

    if not out_rows:
        print(f"  Skipping '{ws.title}': no data rows found")
        return

    os.makedirs(output_dir, exist_ok=True)
    print(f"  {ws.title} → {filepath} ({len(out_rows)} teams)")
    for note in notes:
        print(note)
    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADER)
        writer.writerows(out_rows)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(script_dir, "temp_mlts.xlsx")
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(script_dir, "mlts_results")

    if not os.path.exists(xlsx_path):
        sys.exit(f"File not found: {xlsx_path}")

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    alt_map = build_alternate_map(wb)
    if alt_map:
        print(f"  Alternates from Points sheet: {', '.join(f'{a} → {c2}' for (_, a), (_, c2, _) in alt_map.items())}")

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        # Skip sheets with no real data (placeholder rows have None anglers)
        data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if not is_blank(r[1]) and not is_blank(r[6])]
        if not data_rows:
            print(f"  Skipping '{sheet_name}': no data yet")
            continue
        convert_sheet(ws, output_dir, alt_map)

    print("Done.")


if __name__ == "__main__":
    main()
